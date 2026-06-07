#!/usr/bin/env python3
"""Batch black-box agentic medical hallucination corrector (Serper). Head-to-head vs REVISE."""
import argparse, json, os, sys
import requests
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SERPER_API_KEY = os.getenv("SERPER_API_KEY")
client = OpenAI()

SYSTEM_PROMPT = (
    "You are a careful medical fact-checker. You are given a medical statement that may contain "
    "hallucinated or incorrect claims, and the question it should answer.\n\n"
    "Use the web_search tool to find authoritative sources (PubMed, NIH, peer-reviewed journals). "
    "Do NOT rely on unverified prior knowledge -- verify with search. Pay close attention to results "
    "whose title or snippet directly matches the question; trust a directly-matching study over generic pages. "
    "You may search multiple times.\n\n"
    "When confident, output a FINAL corrected statement consistent with the evidence, concise and on-topic. "
    "If already correct, restate it. If evidence is insufficient, say so rather than guessing.\n"
    "Begin the final answer with 'CORRECTED:' on its own line."
)
TOOLS = [{"type": "function", "function": {
    "name": "web_search",
    "description": "Search the web (Google via Serper) for authoritative medical sources.",
    "parameters": {"type": "object",
        "properties": {"query": {"type": "string", "description": "The search query."}},
        "required": ["query"]}}}]

def web_search(query, k=5):
    if not SERPER_API_KEY:
        return "ERROR: SERPER_API_KEY not set.", []
    try:
        resp = requests.post("https://google.serper.dev/search",
            headers={"X-API-KEY": SERPER_API_KEY, "Content-Type": "application/json"},
            json={"q": query, "num": k}, timeout=20)
        resp.raise_for_status()
        rows = resp.json().get("organic", []) or []
        if not rows:
            return "No results.", []
        text = "\n".join(f"- {r.get('title','')}: {r.get('snippet','')} ({r.get('link','')})" for r in rows[:k])
        links = [r.get("link","") for r in rows[:k] if r.get("link")]
        return text, links
    except Exception as e:
        return f"Search error: {e}", []

def run_agent(paragraph, question, model="gpt-4o-mini", max_steps=6, verbose=True):
    user = f"Statement to check:\n{paragraph}\n\nQuestion it should answer:\n{question}\n"
    messages = [{"role": "system", "content": SYSTEM_PROMPT}, {"role": "user", "content": user}]
    queries, sources, evidence = [], [], []
    step = 0
    for _ in range(max_steps):
        step += 1
        resp = client.chat.completions.create(model=model, messages=messages,
            tools=TOOLS, tool_choice="auto", temperature=0.0)
        msg = resp.choices[0].message
        entry = {"role": "assistant", "content": msg.content or ""}
        if msg.tool_calls:
            entry["tool_calls"] = [{"id": tc.id, "type": "function",
                "function": {"name": tc.function.name, "arguments": tc.function.arguments}} for tc in msg.tool_calls]
        messages.append(entry)
        if msg.tool_calls:
            for tc in msg.tool_calls:
                try:
                    args = json.loads(tc.function.arguments or "{}")
                except json.JSONDecodeError:
                    args = {}
                q = args.get("query","")
                queries.append(q)
                text, links = web_search(q)
                sources.extend(links)
                evidence.append({"step": step, "query": q, "snippets": text})
                if verbose:
                    print(f"    [step {step}] SEARCH: {q}", file=sys.stderr)
                    for ln in text.splitlines():
                        print(f"        {ln[:160]}", file=sys.stderr)
                messages.append({"role": "tool", "tool_call_id": tc.id, "content": text})
            continue
        ans = msg.content or ""
        corrected = ans.split("CORRECTED:")[-1].strip() if "CORRECTED:" in ans else ans.strip()
        return corrected, queries, list(dict.fromkeys(sources)), evidence
    return "(stopped: max_steps reached)", queries, list(dict.fromkeys(sources)), evidence

def export_excel(results, path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Agent Results"
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", start_color="1F4E79")
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wa = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["#","Question","Hallucinated Claim","Ground Truth","Agent Corrected Answer","Search Queries"]
    widths = [5,26,40,40,40,30]
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(1,c,h); cell.font=hf; cell.fill=hfill; cell.alignment=ca; cell.border=bd
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 25
    for i,o in enumerate(results,1):
        row=i+1
        vals=[i,o.get("question_text",""),o.get("claim",""),o.get("long_answer",""),
              o.get("agent_answer","")," | ".join(o.get("agent_queries",[]))]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row,c,v); cell.alignment=ca if c==1 else wa
            cell.font=Font(name="Arial",size=10); cell.border=bd
        ws.row_dimensions[row].height = 90
    wb.save(path)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="medhallu_30.jsonl")
    ap.add_argument("--output", default="agent_30.jsonl")
    ap.add_argument("--model", default="gpt-4o-mini")
    ap.add_argument("--max_steps", type=int, default=6)
    ap.add_argument("--limit", type=int, default=0)
    args = ap.parse_args()
    if not os.path.exists(args.input):
        sys.exit(f"Input not found: {args.input}")
    records = [json.loads(l) for l in open(args.input, encoding="utf-8") if l.strip()]
    if args.limit:
        records = records[:args.limit]
    print(f"Loaded {len(records)} claims from {args.input}. Model={args.model}", file=sys.stderr)
    results = []
    for i, rec in enumerate(records, 1):
        claim = rec.get("claim",""); question = rec.get("question_text","") or rec.get("question","")
        print(f"[{i}/{len(records)}] {claim[:60]}...", file=sys.stderr)
        try:
            answer, queries, sources, evidence = run_agent(claim, question, model=args.model, max_steps=args.max_steps)
        except Exception as e:
            print(f"   ERROR: {e}", file=sys.stderr); answer, queries, sources, evidence = f"(error: {e})", [], [], []
        results.append({**rec, "agent_answer": answer, "agent_queries": queries, "agent_sources": sources, "agent_evidence": evidence})
    with open(args.output, "w", encoding="utf-8") as f:
        for r in results:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    xlsx = args.output.replace(".jsonl", ".xlsx")
    export_excel(results, xlsx)
    print(f"\nWrote {len(results)} -> {args.output} and {xlsx}", file=sys.stderr)

if __name__ == "__main__":
    main()
